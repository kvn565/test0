import os
from django.conf import settings
from django.utils import timezone

from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_LEFT, TA_CENTER

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def creer_dossier(type_export):
    chemin = os.path.join(settings.MEDIA_ROOT, 'rapports', type_export)
    os.makedirs(chemin, exist_ok=True)
    return chemin


def supprimer_ancien_fichier(chemin_complet):
    if os.path.exists(chemin_complet):
        try:
            os.remove(chemin_complet)
        except Exception:
            pass


def _formater_periode(date_debut=None, date_fin=None):
    """✅ Construit la ligne de période lisible pour PDF et Excel"""
    if not date_debut and not date_fin:
        return "Période : toutes les dates"

    def fmt(d):
        """Convertit YYYY-MM-DD en DD/MM/YYYY"""
        try:
            parts = str(d).split('-')
            return f"{parts[2]}/{parts[1]}/{parts[0]}"
        except Exception:
            return str(d)

    if date_debut and date_fin:
        return f"Période : du {fmt(date_debut)} au {fmt(date_fin)}"
    elif date_debut:
        return f"Période : à partir du {fmt(date_debut)}"
    else:
        return f"Période : jusqu'au {fmt(date_fin)}"


def generer_pdf(titre, colonnes, lignes, type_rapport, orientation="landscape",
                date_debut=None, date_fin=None):
    """Génère un PDF en paysage avec métadonnées de période en en-tête"""
    dossier = creer_dossier('pdf')
    nom_fichier = f"{type_rapport}.pdf"
    chemin_complet = os.path.join(dossier, nom_fichier)

    supprimer_ancien_fichier(chemin_complet)

    pagesize = landscape(A4)

    try:
        doc = SimpleDocTemplate(
            chemin_complet,
            pagesize=pagesize,
            rightMargin=30,
            leftMargin=30,
            topMargin=40,
            bottomMargin=30
        )

        styles = getSampleStyleSheet()

        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=9,
            leading=11,
            alignment=TA_LEFT,
            wordWrap='CJK',
            spaceAfter=2,
            allowWidows=0,
            allowOrphans=0,
            splitLongWords=True,
        )

        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontSize=9,
            leading=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold',
            textColor=colors.whitesmoke,
        )

        # ✅ Style sous-titre pour période et date génération
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontSize=10,
            leading=14,
            alignment=TA_LEFT,
            textColor=colors.HexColor('#374151'),
        )

        meta_style = ParagraphStyle(
            'MetaStyle',
            parent=styles['Normal'],
            fontSize=9,
            leading=12,
            alignment=TA_LEFT,
            textColor=colors.HexColor('#6b7280'),
        )

        elements = []

        # ✅ Titre principal
        elements.append(Paragraph(f"<b>{titre}</b>", styles['Title']))

        # ✅ Ligne période
        periode = _formater_periode(date_debut, date_fin)
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(periode, subtitle_style))

        # ✅ Ligne date de génération
        date_gen = timezone.now().strftime('%d/%m/%Y à %H:%M')
        elements.append(Paragraph(f"Généré le : {date_gen}", meta_style))
        elements.append(Spacer(1, 16))

        # Données du tableau
        data = [[]]
        for col in colonnes:
            data[0].append(Paragraph(str(col), header_style))

        for row in lignes:
            new_row = []
            for i, cell in enumerate(row):
                cell_str = str(cell) if cell is not None else '—'
                if i in [2, 3, 7, 8]:
                    new_row.append(Paragraph(cell_str, cell_style))
                else:
                    new_row.append(cell_str)
            data.append(new_row)

        table = Table(
            data,
            colWidths=None,
            repeatRows=1,
            splitByRow=True
        )

        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))

        elements.append(table)
        doc.build(elements)

        return f"rapports/pdf/{nom_fichier}"

    except Exception as e:
        print(f"[ERREUR PDF] {e}")
        raise


def generer_excel(titre, colonnes, lignes, type_rapport,
                  date_debut=None, date_fin=None):
    """Génère un Excel avec métadonnées de période en en-tête"""
    dossier = creer_dossier('excel')
    nom_fichier = f"{type_rapport}.xlsx"
    chemin_complet = os.path.join(dossier, nom_fichier)

    supprimer_ancien_fichier(chemin_complet)

    try:
        wb = Workbook()
        ws = wb.active
        
        # ✅ Nettoyage du titre pour Excel (supprime les caractères interdits)
        titre_nettoye = titre.replace('/', '-').replace('\\', '-').replace(':', '').replace('?', '').replace('*', '').replace('[', '').replace(']', '')
        ws.title = titre_nettoye[:31]   # Maximum 31 caractères pour Excel

        # Styles (le reste du code reste inchangé)
        titre_font    = Font(bold=True, size=13, color="374151")
        periode_font  = Font(size=10, color="374151")
        meta_font     = Font(size=9, italic=True, color="6b7280")
        header_font   = Font(bold=True, color="FFFFFF")
        header_fill   = PatternFill(start_color="374151", end_color="374151", fill_type="solid")

        nb_cols = len(colonnes)

        # Ligne 1 — Titre
        ws.append([titre])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_cols)
        ws.cell(row=1, column=1).font = titre_font
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='left')
        ws.row_dimensions[1].height = 20

        # Ligne 2 — Période
        periode = _formater_periode(date_debut, date_fin)
        ws.append([periode])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nb_cols)
        ws.cell(row=2, column=1).font = periode_font
        ws.cell(row=2, column=1).alignment = Alignment(horizontal='left')

        # Ligne 3 — Date de génération
        date_gen = timezone.now().strftime('%d/%m/%Y à %H:%M')
        ws.append([f"Généré le : {date_gen}"])
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=nb_cols)
        ws.cell(row=3, column=1).font = meta_font
        ws.cell(row=3, column=1).alignment = Alignment(horizontal='left')

        # Ligne 4 — vide
        ws.append([])

        # Ligne 5 — En-têtes
        ws.append(colonnes)
        for cell in ws[5]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Données
        for ligne in lignes:
            ws.append(ligne)

        # Ajustement automatique des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value or "")) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 3, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(chemin_complet)
        return f"rapports/excel/{nom_fichier}"

    except Exception as e:
        print(f"[ERREUR EXCEL] {e}")
        raise